#!/usr/bin/env python3
"""Build month-level category and product profit rows for the AGP dashboard.

The product source is the confirmed EZ-admin shipping SKU match.  Each matched
order's net payment is allocated to its shipped SKUs by COGS weight so product
revenue reconciles exactly to covered order revenue without inventing a selling
price for component SKUs.  Advertising and fixed costs are intentionally not
allocated at product level.
"""

from __future__ import annotations

import argparse
from collections import defaultdict
from datetime import date
from decimal import Decimal, ROUND_HALF_UP
import json
import os
from pathlib import Path
import re

import openpyxl
import psycopg2


SOURCE_SYSTEMS = ("ga4_self_store", "naver_commerce")
COST_ALIASES = {
    "[추가]아이스팩": "아이스팩 추가",
}
BALANCY_SET_COST_SKUS = {
    "밸런시 마라 280g",
    "밸런시 시그니처 280g",
    "밸런시 커리 280g",
    "밸런시 토마토 280g",
}
CATEGORY_ORDER = [
    "단백밥",
    "순수단백",
    "닭가슴살",
    "함박스테이크",
    "밸런시",
    "소스",
    "통밀빵",
    "땅콩버터",
    "알룰로스",
    "부가옵션",
    "기타",
]


def as_decimal(value) -> Decimal:
    if value is None or value == "":
        return Decimal("0")
    return Decimal(str(value))


def json_number(value: Decimal):
    rounded = value.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    if rounded == rounded.to_integral_value():
        return int(rounded)
    return float(rounded)


def next_month(month: str) -> str:
    year, number = map(int, month.split("-"))
    if number == 12:
        return f"{year + 1:04d}-01-01"
    return f"{year:04d}-{number + 1:02d}-01"


def category_for_sku(sku_name: str) -> str:
    name = sku_name.strip()
    if "아이스팩" in name or "배송" in name:
        return "부가옵션"
    if "밸런시" in name or "곡물볶음밥" in name:
        return "밸런시"
    if "소스" in name or "드레싱" in name:
        return "소스"
    if "통밀" in name:
        return "통밀빵"
    if "땅콩버터" in name:
        return "땅콩버터"
    if "알룰로스" in name:
        return "알룰로스"
    if "단백밥" in name or "도시락" in name:
        return "단백밥"
    if "순수단백" in name:
        return "순수단백"
    if "닭가슴살" in name:
        return "닭가슴살"
    if "함박스테이크" in name:
        return "함박스테이크"
    return "기타"


def load_costs(path: Path) -> dict[str, Decimal]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    result: dict[str, Decimal] = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            break
        name = str(row[1] or "").strip()
        if name:
            result[name] = as_decimal(row[2])
    return result


def unit_cost_for_source(source_system: str, sku_name: str, stored_cost: Decimal) -> Decimal:
    # The self-store matcher expands a 1.12kg Balancy set into four 280g units,
    # while Naver keeps one set row.  This mirrors the two channel calculators.
    if source_system == "ga4_self_store" and sku_name in BALANCY_SET_COST_SKUS:
        return stored_cost / Decimal("4")
    return stored_cost


def allocate_order(revenue: Decimal, lines: list[tuple[str, Decimal, Decimal]]) -> list[dict]:
    """Allocate one order's revenue by extended COGS, preserving the exact sum."""
    if not lines:
        raise ValueError("order has no matched SKU lines")
    total_cogs = sum((qty * unit_cost for _name, qty, unit_cost in lines), Decimal("0"))
    if total_cogs <= 0:
        raise ValueError("order has no positive matched COGS")

    allocations = []
    assigned = Decimal("0")
    for index, (name, qty, unit_cost) in enumerate(lines):
        line_cogs = qty * unit_cost
        if index == len(lines) - 1:
            allocated_revenue = revenue - assigned
        else:
            allocated_revenue = (revenue * line_cogs / total_cogs).quantize(
                Decimal("0.01"), rounding=ROUND_HALF_UP
            )
            assigned += allocated_revenue
        allocations.append(
            {
                "name": name,
                "qty": qty,
                "revenue": allocated_revenue,
                "cogs": line_cogs,
                "profit": allocated_revenue - line_cogs,
            }
        )
    return allocations


def fetch_month_data(conn, month: str) -> dict:
    start = f"{month}-01"
    end = next_month(month)
    with conn.cursor() as cur:
        cur.execute(
            """
            select source_system, source_order_id,
                   coalesce(net_payment_amount, payment_amount, 0) as revenue
            from fact_order
            where is_valid_purchase
              and source_system = any(%s)
              and paid_datetime >= %s::date
              and paid_datetime < %s::date
            order by source_system, source_order_id
            """,
            (list(SOURCE_SYSTEMS), start, end),
        )
        orders = {(row[0], row[1]): as_decimal(row[2]) for row in cur.fetchall()}

        cur.execute(
            """
            with costed_matching as (
                select sem.source_system, sem.source_order_id, sem.matched_sku_name,
                       sem.matched_qty,
                       case
                         when sem.source_system = 'ga4_self_store'
                          and sem.matched_sku_name = any(%s)
                         then coalesce(cm.cogs, 0) / 4
                         else coalesce(cm.cogs, 0)
                       end as unit_cost
                from stg_ezadmin_order_match sem
                join fact_order fo
                  on fo.source_system = sem.source_system
                 and fo.source_order_id = sem.source_order_id
                left join lateral (
                    select cost.cogs
                    from stg_cost_master_sku cost
                    where cost.normalized_sku_name = regexp_replace(
                              lower(sem.matched_sku_name), '[^a-z0-9가-힣]+', '', 'g'
                          )
                      and coalesce(cost.effective_start_date, date '1900-01-01') <= fo.paid_datetime::date
                    order by coalesce(cost.effective_start_date, date '1900-01-01') desc
                    limit 1
                ) cm on true
                where sem.source_system = any(%s)
                  and sem.report_date >= %s::date
                  and sem.report_date < %s::date
                  and sem.match_status = 'matched'
                  and coalesce(sem.matched_sku_name, '') <> ''
                  and sem.matched_qty > 0
            )
            select source_system, source_order_id, matched_sku_name,
                   sum(matched_qty) as matched_qty, max(unit_cost) as unit_cost
            from costed_matching
            group by 1, 2, 3
            order by 1, 2, 3
            """,
            (list(BALANCY_SET_COST_SKUS), list(SOURCE_SYSTEMS), start, end),
        )
        matching = defaultdict(list)
        for source, order_id, sku_name, qty, unit_cost in cur.fetchall():
            matching[(source, order_id)].append(
                (str(sku_name).strip(), as_decimal(qty), as_decimal(unit_cost))
            )

        cur.execute(
            """
            select coalesce(sum(total_order_count), 0),
                   coalesce(sum(total_payment_amount), 0)
            from mart_daily_profit_gauge
            where report_date >= %s::date and report_date < %s::date
            """,
            (start, end),
        )
        gauge_orders, gauge_revenue = cur.fetchone()

        cur.execute(
            """
            select
              coalesce((select sum(total_cost) from imweb_profit_daily_summary
                        where source = 'ga4' and date_key::date >= %s::date and date_key::date < %s::date), 0)
              +
              coalesce((select sum(cogs) from vw_naver_commerce_profit_daily
                        where report_date >= %s::date and report_date < %s::date), 0)
            """,
            (start, end, start, end),
        )
        official_channel_cogs = as_decimal(cur.fetchone()[0])

    return {
        "orders": orders,
        "matching": matching,
        "gauge_orders": int(gauge_orders or 0),
        "gauge_revenue": as_decimal(gauge_revenue),
        "official_channel_cogs": official_channel_cogs,
    }


def build_rows(data: dict, costs: dict[str, Decimal], product_limit: int = 15) -> tuple[list, list, dict]:
    orders = data["orders"]
    matching = data["matching"]
    product_totals = defaultdict(lambda: {"qty": Decimal("0"), "revenue": Decimal("0"), "cogs": Decimal("0"), "profit": Decimal("0")})
    missing_cost_skus = set()
    matched_orders = 0
    matched_revenue = Decimal("0")

    for order_key, revenue in orders.items():
        raw_lines = matching.get(order_key, [])
        if not raw_lines:
            continue
        source_system = order_key[0]
        lines = []
        for raw_name, qty, authoritative_unit_cost in raw_lines:
            name = COST_ALIASES.get(raw_name, raw_name)
            if authoritative_unit_cost > 0:
                unit_cost = authoritative_unit_cost
            else:
                stored_cost = costs.get(name, Decimal("0"))
                unit_cost = unit_cost_for_source(source_system, name, stored_cost)
            if unit_cost <= 0:
                missing_cost_skus.add(name)
            lines.append((name, qty, unit_cost))
        if any(unit_cost <= 0 for _name, _qty, unit_cost in lines):
            continue

        matched_orders += 1
        matched_revenue += revenue
        for row in allocate_order(revenue, lines):
            bucket = product_totals[row["name"]]
            for field in ("qty", "revenue", "cogs", "profit"):
                bucket[field] += row[field]

    if missing_cost_skus:
        raise RuntimeError("원가표에 없는 출고 SKU: " + ", ".join(sorted(missing_cost_skus)))

    all_products = [
        {
            "name": name,
            "qty": json_number(values["qty"]),
            "revenue": json_number(values["revenue"]),
            "cogs": json_number(values["cogs"]),
            "profit": json_number(values["profit"]),
        }
        for name, values in product_totals.items()
    ]
    all_products.sort(key=lambda row: (-float(row["profit"]), -float(row["revenue"]), row["name"]))

    category_totals = defaultdict(lambda: {"qty": Decimal("0"), "revenue": Decimal("0"), "cogs": Decimal("0"), "profit": Decimal("0")})
    for name, values in product_totals.items():
        bucket = category_totals[category_for_sku(name)]
        for field in ("qty", "revenue", "cogs", "profit"):
            bucket[field] += values[field]
    categories = [
        {
            "name": name,
            "qty": json_number(values["qty"]),
            "revenue": json_number(values["revenue"]),
            "cogs": json_number(values["cogs"]),
            "profit": json_number(values["profit"]),
        }
        for name, values in category_totals.items()
    ]
    categories.sort(key=lambda row: CATEGORY_ORDER.index(row["name"]) if row["name"] in CATEGORY_ORDER else 99)

    fact_orders = len(orders)
    fact_revenue = sum(orders.values(), Decimal("0"))
    gauge_orders = data["gauge_orders"]
    gauge_revenue = data["gauge_revenue"]
    shipping_cogs = sum((values["cogs"] for values in product_totals.values()), Decimal("0"))
    official_channel_cogs = data["official_channel_cogs"]
    meta = {
        "factOrders": fact_orders,
        "matchedOrders": matched_orders,
        "unmatchedOrders": fact_orders - matched_orders,
        "matchedRevenue": json_number(matched_revenue),
        "unmatchedRevenue": json_number(fact_revenue - matched_revenue),
        "gaugeOrders": gauge_orders,
        "gaugeRevenue": json_number(gauge_revenue),
        "factOrderGapVsGauge": gauge_orders - fact_orders,
        "factRevenueGapVsGauge": json_number(gauge_revenue - fact_revenue),
        "shippingCogs": json_number(shipping_cogs),
        "officialChannelCogs": json_number(official_channel_cogs),
        "cogsGapVsDailyProfit": json_number(shipping_cogs - official_channel_cogs),
        "coveragePercent": float((matched_revenue / gauge_revenue * 100).quantize(Decimal("0.1"))) if gauge_revenue else 0.0,
        "allocationMethod": "주문 결제액을 출고 SKU 원가 비중으로 배분",
        "productRows": len(all_products),
        "categoryRows": len(categories),
    }

    product_revenue = sum((as_decimal(row["revenue"]) for row in all_products), Decimal("0"))
    category_revenue = sum((as_decimal(row["revenue"]) for row in categories), Decimal("0"))
    if abs(product_revenue - matched_revenue) > Decimal("0.02"):
        raise RuntimeError(f"제품 매출 대사 실패: {product_revenue} != {matched_revenue}")
    if abs(category_revenue - matched_revenue) > Decimal("0.02"):
        raise RuntimeError(f"카테고리 매출 대사 실패: {category_revenue} != {matched_revenue}")
    return categories, all_products[:product_limit], meta


def render_entry(month: str, value, indent: str = "      ") -> str:
    rendered = json.dumps(value, ensure_ascii=False, indent=2)
    lines = rendered.splitlines()
    return f'{indent}"{month}": ' + (lines[0] if len(lines) == 1 else lines[0] + "\n" + "\n".join(indent + line for line in lines[1:]))


def upsert_const_month(text: str, var_name: str, month: str, value) -> str:
    marker = f"const {var_name} = {{"
    start = text.find(marker)
    if start < 0:
        raise RuntimeError(f"{var_name} 상수를 찾지 못했습니다")
    end = text.find("\n    };", start)
    if end < 0:
        raise RuntimeError(f"{var_name} 객체 끝을 찾지 못했습니다")
    block = text[start:end]
    entry = render_entry(month, value)
    pattern = re.compile(r'      "' + re.escape(month) + r'": \[.*?\n      \](?=,?\n      "|\n?$)', re.DOTALL)
    if pattern.search(block):
        block = pattern.sub(entry, block, count=1)
    else:
        block = block.rstrip() + ",\n" + entry
    return text[:start] + block + text[end:]


def upsert_meta_month(text: str, month: str, meta: dict) -> str:
    var_name = "productProfitMetaByMonth"
    if f"const {var_name} = {{" not in text:
        product_marker = "const productRowsByMonth = {"
        start = text.find(product_marker)
        end = text.find("\n    };", start)
        if start < 0 or end < 0:
            raise RuntimeError("productRowsByMonth 위치를 찾지 못했습니다")
        insert_at = end + len("\n    };")
        block = "\n    const productProfitMetaByMonth = {\n" + render_entry(month, meta) + "\n    };"
        return text[:insert_at] + block + text[insert_at:]

    marker = f"const {var_name} = {{"
    start = text.find(marker)
    end = text.find("\n    };", start)
    block = text[start:end]
    entry = render_entry(month, meta)
    pattern = re.compile(r'      "' + re.escape(month) + r'": \{.*?\n      \}(?=,?\n      "|\n?$)', re.DOTALL)
    if pattern.search(block):
        block = pattern.sub(entry, block, count=1)
    else:
        block = block.rstrip() + ",\n" + entry
    return text[:start] + block + text[end:]


def update_html(path: Path, month: str, categories: list, products: list, meta: dict, dry_run: bool) -> None:
    text = path.read_text(encoding="utf-8")
    text = upsert_const_month(text, "categoryRowsByMonth", month, categories)
    text = upsert_const_month(text, "productRowsByMonth", month, products)
    text = upsert_meta_month(text, month, meta)
    if dry_run:
        print(json.dumps({"month": month, "category_rows": len(categories), "product_rows": len(products), "meta": meta}, ensure_ascii=False))
        return
    path.write_text(text, encoding="utf-8")
    print(json.dumps({"updated": str(path), "month": month, "category_rows": len(categories), "product_rows": len(products), "meta": meta}, ensure_ascii=False))


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--month", required=True, help="YYYY-MM")
    parser.add_argument("--html", default=str(Path(__file__).resolve().parent.parent / "index.html"))
    parser.add_argument("--cost-master", default=str(Path.home() / "Desktop" / "원가관리.xlsx"))
    parser.add_argument("--product-limit", type=int, default=15)
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    if not re.fullmatch(r"\d{4}-\d{2}", args.month):
        raise SystemExit("--month 형식은 YYYY-MM 입니다")
    database_url = os.environ.get("DATABASE_URL", "")
    if not database_url:
        raise SystemExit("DATABASE_URL 환경변수가 필요합니다")
    cost_path = Path(args.cost_master).expanduser()
    if not cost_path.exists():
        raise SystemExit(f"원가표가 없습니다: {cost_path}")

    costs = load_costs(cost_path)
    with psycopg2.connect(database_url) as conn:
        conn.set_session(readonly=True)
        data = fetch_month_data(conn, args.month)
    categories, products, meta = build_rows(data, costs, product_limit=args.product_limit)
    if not categories or not products:
        raise SystemExit(f"{args.month} 제품 손익 행이 비어 있습니다")
    update_html(Path(args.html), args.month, categories, products, meta, args.dry_run)


if __name__ == "__main__":
    main()
